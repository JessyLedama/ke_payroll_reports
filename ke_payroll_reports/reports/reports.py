# -*- coding: utf-8 -*-
# License LGPL-3.0 or later (https://opensource.org/licenses/LGPL-3.0).
#
# This software and associated files (the "Software") may only be used (executed,
# modified, executed after modifications) if you have purchased a valid license
# from the authors, typically via Odoo Apps, or if you have received a written
# agreement from the authors of the Software (see the COPYRIGHT section below).
#
# You may develop Odoo modules that use the Software as a library (typically
# by depending on it, importing it and using its resources), but without copying
# any source code or material from the Software. You may distribute those
# modules under the license of your choice, provided that this license is
# compatible with the terms of the Odoo Proprietary License (For example:
# LGPL, MIT, or proprietary licenses similar to this one).
#
# It is forbidden to publish, distribute, sublicense, or sell copies of the Software
# or modified copies of the Software.
#
# The above copyright notice and this permission notice must be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT.
# IN NO EVENT SHALL THE AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM,
# DAMAGES OR OTHER LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE,
# ARISING FROM, OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER
# DEALINGS IN THE SOFTWARE.
#
######### COPYRIGHT#####
# © 2016 Bernard K Too<bernard.too@optima.co.ke>
import base64
import enum
import logging
import os
import re
import string
import tempfile

from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter

from odoo import _, api, fields, models
from odoo.exceptions import AccessError, UserError, ValidationError

_logger = logging.getLogger(__name__)
try:
    import openpyxl
except ImportError:
    msg = _(
        'Install python module "openpyxl" in order to create Excel documents')
    raise ValidationError(msg)
try:
    import csv
except ImportError:
    msg = _('Install python module "csv" in order to generate CSV')
    raise ValidationError(msg)

MID_FONT = Font(name='Arial', bold=True, size=12)
NORMAL_FONT = Font(name='Arial', bold=False, size=10)
# Paper size
# PAPERSIZE_LETTER = '1'
# PAPERSIZE_LETTER_SMALL = '2'
# PAPERSIZE_TABLOID = '3'
# PAPERSIZE_LEDGER = '4'
# PAPERSIZE_LEGAL = '5'
# PAPERSIZE_STATEMENT = '6'
# PAPERSIZE_EXECUTIVE = '7'
# PAPERSIZE_A3 = '8'
# PAPERSIZE_A4 = '9'
# PAPERSIZE_A4_SMALL = '10'
# PAPERSIZE_A5 = '11'

# Page orientation
# ORIENTATION_PORTRAIT = 'portrait'
# ORIENTATION_LANDSCAPE = 'landscape'


class PayslipReports(models.Model):
    _inherit = "hr.payslip"
    _name = "hr.payslip"

    def GetPayslipExcel(self):
        for rec in self:
            filename_slip = 'Payslip_' + re.sub(
                '[^A-Za-z0-9]+', '_',
                rec.number) + '-' + fields.Datetime.context_timestamp(
                    self, fields.Datetime.now()).strftime(
                        '%Y_%m_%d-%H%M%S') + '.xlsx'
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.fitToHeight = 0
            ws.page_setup.fitToWidth = 1
            # t = 0
            ############## STYLES##################
            # Number format
            nf = '#,##0.00'
            # TableHeaderPattern
            th = PatternFill(
                start_color=re.sub('[#]+', '', rec.theme_color)
                or 'FFFFFF' + '00',
                end_color=re.sub('[#]+', '', rec.theme_color)
                or 'FFFFFF' + '00',
                fill_type='solid',
            )
            # ThemeTextFont
            tt = Font(
                color=re.sub('[#]+', '', rec.theme_txt_color)
                or 'FFFFFF' + '00',
                size=int(rec.body_font) + 2 or None,
                name=rec.font_family or None,
                bold=True,
            )
            # BodyTextFont
            bt = Font(
                color=re.sub('[#]+', '', rec.text_color) or '000000' + '00',
                size=rec.body_font or None,
                name=rec.font_family or None,
            )
            # BodyTextBoldFont
            bb = Font(
                color=re.sub('[#]+', '', rec.text_color) or '000000' + '00',
                size=rec.body_font or None,
                name=rec.font_family or None,
                bold=True,
            )
            # CompanyName Font
            ct = Font(
                color=re.sub('[#]+', '', rec.name_color) or '000000' + '00',
                size=int(rec.header_font) + 2 or None,
                name=rec.font_family or None,
                bold=True,
            )
            # HeaderText Font
            ht = Font(
                color=re.sub('[#]+', '', rec.text_color) or '000000' + '00',
                size=rec.header_font or None,
                name=rec.font_family or None,
            )
            # Alignment
            al1 = Alignment(horizontal="left", vertical="center")
            al2 = Alignment(horizontal="right", vertical="center")
            # Company Address
            ws['A1'] = rec.company_id.name or None
            ws.merge_cells('A1:C1')
            ws['A1'].font = ct
            ws['A2'] = rec.company_id.street or None
            ws.merge_cells('A2:C2')
            ws['A2'].font = ht
            ws['A3'] = rec.company_id.street2 or None
            ws.merge_cells('A3:C3')
            ws['A3'].font = ht
            ws['A4'] = (rec.company_id.city
                        or '') + (rec.company_id.zip and
                                  (', ' + rec.company_id.zip) or '')
            ws.merge_cells('A4:C4')
            ws['A4'].font = ht
            ws['A5'] = rec.company_id.phone and (
                'Tel: ' + rec.company_id.phone) or 'Tel:'
            ws.merge_cells('A5:C5')
            ws['A5'].font = ht
            # ws['A6'] = rec.company_id.fax and (
            #    'Fax: ' + rec.company_id.fax) or 'Fax:'
            ws.merge_cells('A6:C6')
            ws['A6'].font = ht
            rec.env['hr.ke'].style_range(ws, 'A1:A6', alignment=al1)
            # Company Logo
            logo_fd, logo_path = tempfile.mkstemp(suffix='.png',
                                                  prefix='logo.tmp.')
            with open(logo_path, "wb") as logo:
                logo.write(base64.decodestring(rec.company_id.logo))
                logo.close()
            img = openpyxl.drawing.image.Image(logo_path)
            # rec.env['hr.ke'].delete_tempfile(logo_path)
            ws.add_image(img, 'D1')
            ws.merge_cells('D1:F6')
            rec.env['hr.ke'].style_range(ws, 'D1:D6', alignment=al2)

            # Emloyee Data
            ws.merge_cells('B9:B13')
            ws['A8'] = 'EMP. NAME:'
            ws['B8'] = rec.employee_id.name or None
            ws['A9'] = 'EMP. ADDRESS:'
            ws['B9'] = (
                (rec.employee_id.address_home_id.street or '') + '\n' +
                (rec.employee_id.address_home_id.street2 or '') + '\n' +
                (rec.employee_id.address_home_id.city or '') + ' ' +
                (rec.employee_id.address_home_id.state_id.code or '') + ' ' +
                (rec.employee_id.address_home_id.zip or '') + '\n' +
                (rec.employee_id.address_home_id.country_id.name or ''))

            ws['C8'] = 'DEPT:'
            ws['D8'] = rec.employee_id.department_id.name or None
            ws['C9'] = 'DESIGNATION:'
            ws['D9'] = rec.employee_id.job_id.name or None
            ws['C10'] = 'ID NO.:'
            ws['D10'] = rec.employee_id.identification_id or None
            ws['C11'] = 'EMAIL:'
            ws['D11'] = rec.employee_id.work_email or None
            ws['C12'] = 'DATE EMPLOYED:'
            ws['D12'] = rec.contract_id.date_start or None
            ws['C13'] = 'MARITAL STATUS:'
            ws['D13'] = rec.employee_id.marital or None

            ws['E8'] = 'REF:'
            ws['F8'] = rec.number or None
            ws['E9'] = 'PERIOD START:'
            ws['F9'] = rec.date_from or None
            ws['E10'] = 'PERIOD END:'
            ws['F10'] = rec.date_to or None
            ws['E11'] = 'BANK ACCOUNT'
            ws['F11'] = rec.employee_id.bank_account_id.acc_number or None
            ws['E12'] = 'BANK NAME:'
            ws['F12'] = rec.employee_id.bank_account_id.bank_id.name or None
            ws['E13'] = 'BANK CODE:'
            ws['F13'] = rec.employee_id.bank_account_id.bank_id.bic or None

            rec.env['hr.ke'].style_range(ws, 'A8:F13', font=bt, alignment=al1)
            rec.env['hr.ke'].style_range(ws, 'A8:A13', font=bb)
            rec.env['hr.ke'].style_range(ws, 'C8:C13', font=bb)
            rec.env['hr.ke'].style_range(ws, 'E8:E13', font=bb)
            ws['B9'].alignment = Alignment(wrapText=True)

            fr = 16
            cols = ['ITEM', 'CODE', 'NAME', 'QTY', 'AMOUNT', 'TOTAL']
            for k, x in enumerate(string.ascii_uppercase[0:6]):  # 'ABCDEF'
                align = Alignment(
                    horizontal='right' if cols[k] in
                    ['QTY/RATE', 'AMOUNT', 'TOTAL'] else 'general')
                ws[x + str(fr - 1)] = cols[k]
                ws[x + str(fr - 1)].fill = th
                ws[x + str(fr - 1)].font = tt
                ws[x + str(fr - 1)].alignment = align
            for key, line in enumerate(
                    [x for x in rec.line_ids if x.appears_on_payslip]):
                ws['A' + str(fr + key)] = str(key + 1)
                ws['A' + str(fr + key)].font = bt
                ws['B' + str(fr + key)] = line.code
                ws['B' + str(fr + key)].font = bt
                ws['C' + str(fr + key)] = line.name
                ws['C' + str(fr + key)].font = bt
                ws['D' + str(fr + key)] = line.quantity
                ws['D' + str(fr + key)].number_format = nf
                ws['D' + str(fr + key)].font = bt
                ws['E' + str(fr + key)] = line.amount
                ws['E' + str(fr + key)].number_format = nf
                ws['E' + str(fr + key)].font = bt
                ws['F' + str(fr + key)] = line.total
                ws['F' + str(fr + key)].number_format = nf
                ws['F' + str(fr + key)].font = bt
            xls_path = rec.env['hr.ke'].create_xls()
            # raise ValidationError(xls_path)
            wb.save(xls_path)
            rec.env['hr.ke'].save_attachment(filename_slip, xls_path,
                                             self._name, rec.id)


class PayrollReports(models.Model):
    _name = 'hr.payslip.run'
    _inherit = ['hr.payslip.run', 'mail.thread', 'mail.activity.mixin']

    def BatchConfirmPayslip(self):
        """
        This  method will confirm  payslips and possible generate accounting
        entries for those salary rules that have accounting settings configured
        """
        for batch in self:
            for slip in batch.slip_ids:
                if not slip.line_ids:
                    slip.compute_sheet()
            # get all empty slips
            empty_slips = batch.slip_ids.filtered(
                lambda x: not x.line_ids or not x.
                details_by_salary_rule_category)
            # raise warning if some slips are empty
            if empty_slips:
                msg = _("Missing payslip lines:")
                for num, slip in enumerate(empty_slips):
                    msg += "\n %s. %s" % (num + 1, slip.name)
                raise UserError(msg)
            for slip in batch.slip_ids:
                # only confirm those that have not been previosly confirmed
                if slip.state in ['draft']:
                    slip.action_payslip_done()

    def helb_report(self):
        cols = ['ID NUMBER', 'NAMES', 'STAFF NUMBER', 'HELB AMOUNT']
        slips = self.slip_ids.filtered(
            lambda s: s.line_ids.filtered(lambda l: l.code == 'P107'))
        filename = f'{self.name} HELB REPORT.xlsx'
        fr = 1
        wb = openpyxl.Workbook()
        ws = wb.active
        for index, col in enumerate(cols, start=1):
            letter = get_column_letter(index)
            ws[f'{letter}{fr}'] = col
            ws[f'{letter}{fr}'].font = MID_FONT
        fr += 1

        for slip in slips:
            ws[f'A{fr}'] = slip.employee_id.identification_id or ''
            ws[f'B{fr}'] = slip.employee_id.name or ''
            ws[f'C{fr}'] = slip.employee_id.employee_no or ''
            ws[f'D{fr}'] = slip.line_ids.filtered(
                lambda l: l.code == 'P107').total
            fr += 1
        xls_path = self.env['hr.ke'].create_xls()
        wb.save(xls_path)
        self.env['hr.ke'].save_attachment(
            filename, xls_path, self._name, self.id)

    def GetNSSFReturns(self):
        for rec in self:
            if rec.slip_ids:
                filename_nssf = 'NSSF_Returns-' + re.sub(
                    '[^A-Za-z0-9]+', '',
                    rec.name) + '_' + fields.Datetime.context_timestamp(
                        self, fields.Datetime.now()).strftime(
                            '%Y_%m_%d-%H%M%S') + '.xlsx'
                wb = openpyxl.Workbook()
                ws = wb.active
                t = 0
                fr = 13
                tiers_m = {}
                tiers_e = {}
                total_income = 0.0
                total_tier1e = 0.0
                total_tier2e = 0.0
                total_tier3e = 0.0
                total_v1e = 0.0

                total_tier1m = 0.0
                total_tier2m = 0.0
                total_tier3m = 0.0
                total_v1m = 0.0
                # EMPLOYER DETAILS
                ws['A1'] = 'RETURNS TYPE'
                # Regular Employees Returns File only, does not include daily
                # paid workers
                ws['B1'] = '01'
                ws['A2'] = 'EMPLOYER KRA PIN'
                # Employer KRA PIN in the HR Settings
                ws['B2'] = rec.company_id.vat or None
                ws['A3'] = 'EMPLOYER NSSF NUMBER'
                # Employer NSSF Number in the HR/Payroll Settings
                ws['B3'] = rec.company_id.company_nssf_no or None
                ws['A4'] = 'EMPLOYER NAME'
                ws['B4'] = rec.company_id.name or None
                ws['A5'] = 'FISCAL PERIOD'
                # ws['B5'] = datetime.datetime.strptime(
                #    rec.date_end, '%Y-%m-%d').strftime('%m%Y')
                ws['B5'] = fields.Date.from_string(
                    rec.date_end).strftime('%m%Y')
                # DATA HEADERS
                ws['A' + str(fr - 1)] = 'PAYROLL NUMBER'
                ws['B' + str(fr - 1)] = 'SURNAME'
                ws['C' + str(fr - 1)] = 'OTHER NAMES'
                ws['D' + str(fr - 1)] = 'ID NO'
                ws['E' + str(fr - 1)] = 'MEMBER KRA PIN NO'
                ws['F' + str(fr - 1)] = 'NSSF NUMBER'
                ws['G' + str(fr - 1)] = 'CONTRIBUTION TYPE'
                ws['H' + str(fr - 1)] = 'PENSIONABLE INCOME'
                ws['I' + str(fr - 1)] = 'INCOME TYPE'
                ws['J' + str(fr - 1)] = 'MEMBER CONTRIBUTIONS'
                ws['K' + str(fr - 1)] = 'EMPLOYER CONTRIBUTIONS'
                ws['L' + str(fr - 1)] = 'TOTAL CONTRIBUTIONS'
                # DATA ITSELF
                for key, slip in enumerate(rec.slip_ids):
                    if slip.line_ids:
                        income = slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule30').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total  # Pensionable income
                    else:
                        msg = _(
                            'No Payslip Details!\nPlease compute the payslip for %s'
                            % slip.employee_id.name)
                        raise ValidationError(msg)
                    tiers_m['t1'] = slip.line_ids.search(
                        [('salary_rule_id', '=',
                          rec.env.ref('hr_ke.ke_rule46').id),
                         ('slip_id', '=', slip.id)],
                        limit=1).total  # NSSF Tier I contributions
                    tiers_m['t2'] = slip.line_ids.search(
                        [('salary_rule_id', '=',
                          rec.env.ref('hr_ke.ke_rule47').id),
                         ('slip_id', '=', slip.id)],
                        limit=1).total  # NSSF Tier II contributions
                    tiers_m['t3'] = slip.line_ids.search(
                        [('salary_rule_id', '=',
                          rec.env.ref('hr_ke.ke_rule48').id),
                         ('slip_id', '=', slip.id)],
                        limit=1).total  # NSSF Tier III contributions
                    tiers_m['v1'] = slip.line_ids.search(
                        [('salary_rule_id', '=',
                          rec.env.ref('hr_ke.ke_rule49').id),
                         ('slip_id', '=', slip.id)],
                        limit=1).total  # NSSF Voluntary contributions
                    # Employer Contributions
                    tiers_e['t1'] = slip.line_ids.search(
                        [('salary_rule_id', '=',
                          rec.env.ref('hr_ke.ke_rule56').id),
                         ('slip_id', '=', slip.id)],
                        limit=1).total  # NSSF Tier I contributions
                    tiers_e['t2'] = slip.line_ids.search(
                        [('salary_rule_id', '=',
                          rec.env.ref('hr_ke.ke_rule57').id),
                         ('slip_id', '=', slip.id)],
                        limit=1).total  # NSSF Tier II contributions
                    tiers_e['t3'] = slip.line_ids.search(
                        [('salary_rule_id', '=',
                          rec.env.ref('hr_ke.ke_rule58').id),
                         ('slip_id', '=', slip.id)],
                        limit=1).total  # NSSF Tier III contributions
                    tiers_e['v1'] = slip.line_ids.search(
                        [('salary_rule_id', '=',
                          rec.env.ref('hr_ke.ke_rule59').id),
                         ('slip_id', '=', slip.id)],
                        limit=1).total  # NSSF Voluntary contributions
                    # Dermine the TOTALS
                    total_income += income
                    total_tier1e += tiers_e['t1']
                    total_tier2e += tiers_e['t2']
                    total_tier3e += tiers_e['t3']
                    total_v1e += tiers_e['v1']

                    total_tier1m += tiers_m['t1']
                    total_tier2m += tiers_m['t2']
                    total_tier3m += tiers_m['t3']
                    total_v1m += tiers_m['v1']

                    tiers_m = {
                        x: tiers_m[x]
                        for x in tiers_m.keys() if (tiers_m[x] or tiers_e[x])
                    }  # Ignore Zero Contributions from Member
                    # Ignore Zero Contributions from Employer
                    tiers_e = {
                        x: tiers_e[x]
                        for x in tiers_m.keys() if (tiers_e[x] or tiers_m[x])
                    }
                    for k in sorted(tiers_m):
                        # PAYROLL NUMBER
                        ws['A' + str(fr + key +
                                     t)] = slip.employee_id.employee_no or None
                        # SURNAME
                        ws['B' + str(fr + key +
                                     t)] = slip.employee_id.display_name.split(
                                         ' ')[-1] or ''
                        # OTHER NAMES
                        ws['C' + str(fr + key +
                                     t)] = slip.employee_id.display_name.split(
                                         ' ')[0] or ''
                        # ID NO (National id number/Alien registration
                        # no./Passport no.)
                        ws['D' + str(
                            fr + key + t
                        )] = slip.employee_id.identification_id or slip.employee_id.passport_id or None
                        # MEMBER KRA PIN NO
                        ws['E' + str(fr + key +
                                     t)] = slip.employee_id.tax_pin or None
                        # NSSF NUMBER
                        ws['F' +
                           str(fr + key + t)] = slip.employee_id.nssf or None
                        # CONTRIBUTION TYPE (Value range: 101, 102, 103, 104,
                        # 105, 200)
                        ws['G' + str(
                            fr + key + t
                        )] = k == 't1' and 101 or k == 't2' and 102 or k == 't3' and 103 or k == 'v1' and 105 or None
                        # PENSIONABLE INCOME {AS DEFINED BY THE NSSF ACT 20fr }
                        ws['H' +
                           str(fr + key + t)] = k in ['t1', 't2', 't3'
                                                      ] and income or None
                        # INCOME TYPE (Value range: 1, 2, 3, 4) ....'1' becuase
                        # this return is for for monthly paid workers only
                        ws['I' + str(fr + key + t)] = 1
                        # MEMBER CONTRIBUTIONS
                        ws['J' + str(fr + key + t)] = tiers_m[k]
                        # EMPLOYER CONTRIBUTIONS
                        ws['K' + str(fr + key + t)] = tiers_e[k]
                        ws['L' +
                           str(fr + key +
                               t)] = ws['J' + str(fr + key + t)].value + ws[
                                   'K' + str(fr + key + t)].value
                        t += 1
                    t -= 1
                # SUMMARY DETAILS
                ws['A6'] = 'TOTAL INCOME'
                ws['B6'] = total_income
                ws['A7'] = 'TOTAL MEMBER CONTRIBUTIONS'
                ws['B7'] = (total_tier1m + total_tier2m + total_tier3m +
                            total_v1m)
                ws['A8'] = 'TOTAL EMPLOYER CONTRIBUTIONS'
                ws['B8'] = (total_tier1e + total_tier2e + total_tier3e +
                            total_v1e)
                ws['A9'] = 'TOTAL CONTRIBUTIONS'
                ws['B9'] = ws['B7'].value + ws['B8'].value
                ws['A10'] = 'NO OF RECORDS'
                ws['B10'] = t + len(rec.slip_ids)
                xls_path = self.env['hr.ke'].create_xls()
                wb.save(xls_path)
                rec.env['hr.ke'].save_attachment(filename_nssf, xls_path,
                                                 self._name, rec.id)
            else:
                msg = _('No Payslips to process!')
                raise ValidationError(msg)

    def GetNHIFReturns(self):
        for rec in self:
            if rec.slip_ids:
                filename_nhif = 'NHIF_ByProduct-' + re.sub(
                    '[^A-Za-z0-9]+', '',
                    rec.name) + '_' + fields.Datetime.context_timestamp(
                        self, fields.Datetime.now()).strftime(
                            '%Y_%m_%d-%H%M%S') + '.xlsx'
                wb = openpyxl.Workbook()
                ws = wb.active
                fr = 6
                total = 0.0
                nhif = 0.0
                # EMPLOYER DETAILS
                ws['A1'] = 'EMPLOYER CODE'
                # NHIF NO IN THE HR SETTINGS
                ws['B1'] = rec.company_id.company_nssf_no or None
                ws['A2'] = 'EMPLOYER NAME'
                ws['B2'] = rec.company_id.name or None
                ws['A3'] = 'MONTH OF CONTRIBUTION'
                # ws['B3'] = datetime.datetime.strptime(
                #    rec.date_end, '%Y-%m-%d').strftime('%Y-%m')
                ws['B3'] = fields.Date.from_string(
                    rec.date_end).strftime('%Y-%m')
                # DATA HEADERS
                ws['A' + str(fr - 1)] = 'PAYROLL NO'
                ws['B' + str(fr - 1)] = 'LAST NAME'
                ws['C' + str(fr - 1)] = 'FIRST NAME'
                ws['D' + str(fr - 1)] = 'ID NO'
                ws['E' + str(fr - 1)] = 'NHIF NO'
                ws['F' + str(fr - 1)] = 'AMOUNT'
                # DATA ITSELF
                for key, slip in enumerate(rec.slip_ids):
                    if slip.line_ids:
                        nhif = slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule106').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0  # NHIF contributions
                    else:
                        msg = _(
                            'No Payslip Details!\nPlease compute the payslip for %s'
                            % slip.employee_id.name)
                        raise ValidationError(msg)
                    # CELLS
                    # PAYROLL NUMBER
                    name = slip.employee_id.name.strip().split(' ')
                    ws['A' +
                       str(fr + key)] = slip.employee_id.employee_no or None
                    # LAST NAME
                    ws['B' +
                       str(fr + key)] = name[-1] or ''
                    # FIRST NAME
                    ws['C' +
                       str(fr + key)] = name[0] or ''
                    # ID NO (National id number/Alien registration no./Passport
                    # no.)
                    ws['D' + str(
                        fr + key
                    )] = slip.employee_id.identification_id or slip.employee_id.passport_id or None
                    # NHIF Number
                    ws['E' + str(fr + key)] = slip.employee_id.nhif or None
                    ws['F' + str(fr + key)] = nhif  # Amount contributed
                    total += nhif
                # TOTAL
                ws['E' + str(fr + key + 1)] = 'TOTAL'
                ws['F' + str(fr + key + 1)] = total
                xls_path = self.env['hr.ke'].create_xls()
                wb.save(xls_path)
                rec.env['hr.ke'].save_attachment(filename_nhif, xls_path,
                                                 self._name, rec.id)
            else:
                msg = _('No Payslips to process!')
                raise ValidationError(msg)

    def _prepare_payroll_summary(self):
        data = []
        rules = [
                ('BASIC PAY', 'P010'), ('HOUSE ALLOWANCE', 'P017'), ('INCONVENIENCE ALLOWANCE', 'P014'), ('OT1 (Units)', 'P0900'), ('OT1 (Amount)', 'P012'),
                ('OT2 (Units)', 'P0900'), ('OT2 (Amount)', 'P013'), ('LEAVE PAY (Units)', 'P0900'), ('LEAVE PAY (Amount)', 'P015'), ('NOTICE PAY', 'SA15'),
                ('REIMBURSEMENT', 'P016'), ('BONUS', 'P011'), ('ABSENT (Units)', 'P0900'),
                ('ABSENT (Amount)', 'PI01'), ('GROSS PAY', 'P030'), ('PAYE', 'P101'), ('NSSF', 'P055'), ('NHIF', 'P106'), ('NHDL', 'P111'),
                ('HELB', 'P107'), ('SALARY ADVANCE', 'P108'), ('COMPANY LOAN', 'LOANINS'), ('SACCO', 'P109'), ('SURCHARGE/RECOVERY', 'P113'),
                ('NOTICE DEDUCTION', 'P114'), ('TOTAL DEDUCTIONS', 'P115'), ('NET PAY', 'P120')
            ]
        for index, slip in enumerate(self.slip_ids, start=1):
            employee = slip.employee_id
            vals = {
                'SR.': index,
                'PAYROLL NUMBER': employee.employee_no or None,
                'EMPLOYEE NAME': employee.name,
                'BANK NAME': employee.bankBranch or None,
                'BANK BRANCH CODE': employee.bankCode or None,
                'BANK ACCOUNT NUMBER': employee.accountNumber or None
            }
            for rule, code in rules:
                vals[rule] = slip.line_ids.filtered(lambda line: line.code == code).total
            data.append(vals)
        return data
    
    def GetPayrollSummary(self):
        for rec in self:
            filename = f'{self.name} Payroll_Summary-.xlsx'
            wb = openpyxl.Workbook()
            ws = wb.active
            fr = 1
            employee_cols = ['SR.', 'PAYROLL NUMBER', 'EMPLOYEE NAME']
            rules_cols = [
                'BASIC PAY', 'HOUSE ALLOWANCE', 'INCONVENIENCE ALLOWANCE', 'OT1 (Units)', 'OT1 (Amount)',
                'OT2 (Units)', 'OT2 (Amount)', 'LEAVE PAY (Units)', 'LEAVE PAY (Amount)', 'NOTICE PAY', 'REIMBURSEMENT', 'BONUS', 'ABSENT (Units)',
                'ABSENT (Amount)', 'GROSS PAY', 'PAYE', 'NSSF', 'NHIF', 'NHDL', 'HELB', 'SALARY ADVANCE', 'COMPANY LOAN', 'SACCO', 'SURCHARGE/RECOVERY',
                'NOTICE DEDUCTION',	'TOTAL DEDUCTIONS',	'NET PAY'
            ]
            bank_cols =['BANK NAME', 'BANK BRANCH CODE', 'BANK ACCOUNT NUMBER']
            cols = employee_cols + rules_cols + bank_cols
            
            for index, col in enumerate(cols, start=1):
                letter = get_column_letter(index)
                ws[f'{letter}{fr}'] = col
                ws[f'{letter}{fr}'].font = MID_FONT
            fr += 1
            start_fr = fr
            for sumr in self._prepare_payroll_summary():
                for index, col in enumerate(cols, start=1):
                    letter = get_column_letter(index)
                    ws[f'{letter}{fr}'] = sumr[col]
                    ws[f'{letter}{fr}'].font = NORMAL_FONT
                fr += 1
            end_fr = fr - 1
            
            employee_cols_end = len(employee_cols)
            employee_cols_letter = get_column_letter(employee_cols_end)
            ws[f'{employee_cols_letter}{fr}'] = 'TOTALS'
            
            for index, col in enumerate(rules_cols, start = employee_cols_end + 1):
                letter = get_column_letter(index)
                ws[f'{letter}{fr}'] = f'=SUM({letter}{start_fr}:{letter}{end_fr})'
                ws[f'{letter}{fr}'].font = MID_FONT
            
            # Save file as attachment
            xls_path = self.env['hr.ke'].create_xls()
            wb.save(xls_path)
            rec.env['hr.ke'].save_attachment(filename, xls_path, self._name, self.id)

    def GetNetPay(self):
        for rec in self:
            if rec.slip_ids:
                filename_netpay = 'NET_PAY-' + re.sub(
                    '[^A-Za-z0-9]+', '',
                    rec.name) + '_' + fields.Datetime.context_timestamp(
                        self, fields.Datetime.now()).strftime(
                            '%Y_%m_%d-%H%M%S') + '.xlsx'
                wb = openpyxl.Workbook()
                ws = wb.active
                fr = 7  # First row of data
                ws['A1'] = rec.company_id.name
                ws['A2'] = 'PAYROLL SUMMARY'
                ws['B2'] = rec.name
                ws['A3'] = 'DATE FROM'
                ws['B3'] = rec.date_start
                ws['A4'] = 'DATE TO'
                ws['B4'] = rec.date_end
                cols = [
                    'EMPLOYEE NAME', 'ACCOUNT NO', 'BANK NAME', 'BANK BRANCH',
                    'AMOUNT'
                ]
                # DATA HEADERS
                for k, x in enumerate(string.ascii_uppercase[0:5]):  # 'ABCDE'
                    ws[x + str(fr - 1)] = cols[k]

                for key, slip in enumerate(rec.slip_ids):
                    ws['A' + str(fr + key)] = slip.employee_id.name or None
                    ws['B' + str(
                        fr + key
                    )] = slip.employee_id.bank_account_id.acc_number or None
                    ws['C' + str(
                        fr + key
                    )] = slip.employee_id.bank_account_id.bank_id.name or None
                    ws['D' + str(
                        fr + key
                    )] = slip.employee_id.bank_account_id.bank_id.bic or None
                    ws['E' + str(fr + key)] = slip.line_ids.search(
                        [('salary_rule_id', '=',
                          rec.env.ref('hr_ke.ke_rule120').id),
                         ('slip_id', '=', slip.id)],
                        limit=1).total  # Total Net Pay
                # Totals
                t = fr + key + 1  # last row for Totals
                ws['D' + str(t)] = 'TOTAL'
                # Sum using excel 'SUM' formula
                ws['E' + str(t)] = '=SUM(E' + str(fr) + ':E' + str(t - 1) + ')'
                # Save file as attachment
                xls_path = self.env['hr.ke'].create_xls()
                wb.save(xls_path)
                rec.env['hr.ke'].save_attachment(filename_netpay, xls_path,
                                                 self._name, rec.id)
            else:
                msg = _('No Payslips to process!')
                raise ValidationError(msg)

    def GetP10(self):
        for rec in self:
            filename_employee = 'Employees_Details-' + re.sub(
                '[^A-Za-z0-9]+', '',
                rec.name) + '_' + fields.Datetime.context_timestamp(
                    self,
                    fields.Datetime.now()).strftime('%Y_%m_%d-%H%M%S') + '.csv'
            filename_disabled = 'Disabled_Employees_Details-' + re.sub(
                '[^A-Za-z0-9]+', '',
                rec.name) + '_' + fields.Datetime.context_timestamp(
                    self,
                    fields.Datetime.now()).strftime('%Y_%m_%d-%H%M%S') + '.csv'
            filename_car = 'Car_Benefit_Details-' + re.sub(
                '[^A-Za-z0-9]+', '',
                rec.name) + '_' + fields.Datetime.context_timestamp(
                    self,
                    fields.Datetime.now()).strftime('%Y_%m_%d-%H%M%S') + '.csv'
            details_employee = []
            details_disabled = []
            details_cars = []
            for slip in rec.slip_ids:
                if slip.contract_id.car and slip.contract_id.cars:
                    for car in slip.contract_id.cars:
                        data_cars = [
                            slip.employee_id.tax_pin,  # A
                            slip.employee_id.disability
                            and 'C_Disabled_Employees_Dtls'
                            or 'B_Employees_Dtls',  # B
                            car.name,  # C
                            car.make,  # D
                            car.body in ['saloon']
                            and 'Saloon Hatch Backs and Estates'
                            or car.body in ['pickup']
                            and 'Pick Ups, Panel Vans Uncovered'
                            or car.body in ['cruiser'] and
                            # E
                            'Land Rovers/ Cruisers(excludes Range Rovers and vehicles of similar nature)',
                            car.cc_rate,  # F
                            car.cost_type,  # G
                            car.cost_type in ['Hired'] and car.cost_hire
                            or None,  # H
                            car.cost_type in ['Owned'] and car.cost_own
                            or None,  # I
                        ]
                        details_cars.append(data_cars)

                # Normal Employees without disability
                if not slip.employee_id.disability:
                    if not slip.employee_id.tax_pin:
                        raise ValidationError(
                            _('KRA PIN Number for %s is missing!' %
                              slip.employee_id.name))
                    data_employee = [
                        slip.employee_id.tax_pin,  # A
                        slip.employee_id.name,  # B
                        slip.employee_id.resident and 'Resident'
                        or 'Non-Resident',  # C
                        slip.employee_id.emp_type in ['primary']
                        and 'Primary Employee' or 'Secondary Employee',  # D
                        slip.contract_id.wage or 0.0,  # E
                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule17').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # F House Allowance
                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule14').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # G Transport Allowance
                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule15').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,  # H Leave Pay
                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule13').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # I Overtime Allowance
                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule16').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # J Directors Fee
                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule18').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # K Lump Sum Pay
                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule19').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # L Other Allowances
                        None,  # M
                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule38').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # N Value of Car Benefit
                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule37').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # O sum of other benefits (water,elec,telephone..etc)
                        None,  # P Total non cash pay -computed by P10 Form
                        (not slip.contract_id.house
                         and str(slip.employee_id.global_income))
                        or (slip.contract_id.house
                            and slip.contract_id.house_type == 'director'
                            and str(slip.employee_id.global_income)) or None,
                        # Q Global income (non full time service director)
                        (slip.contract_id.house
                         and slip.contract_id.house_type == 'own'
                         and "Employer's Owned House")
                        or (slip.contract_id.house
                            and slip.contract_id.house_type == 'rented'
                            and "Employer's Rented House")
                        or (slip.contract_id.house
                            and slip.contract_id.house_type == 'agric'
                            and "Agriculture Farm")
                        or (slip.contract_id.house
                            and slip.contract_id.house_type == 'director'
                            and "House to Non full time service Director")
                        or (not slip.contract_id.house
                            and "Benefit not given"),  # R
                        (slip.contract_id.house
                         and slip.contract_id.house_type not in ['director']
                         and slip.contract_id.rent)
                        or (slip.contract_id.house
                            and slip.contract_id.house_type in ['director']
                            and '0.0') or
                        # S Rent of House or Its market Value
                        (not slip.contract_id.house and None),
                        None,  # T Computed Rent of house(15% of GrossPay)
                        (slip.contract_id.house
                         and slip.contract_id.house_type not in ['director']
                         and slip.contract_id.rent_recovered)
                        or (slip.contract_id.house
                            and slip.contract_id.house_type in ['director']
                            and '0.0') or
                        # U Rent Recovered from Employee
                        (not slip.contract_id.house and None),
                        None,  # V Net value of Housing
                        None,  # W  Total Gross Pay
                        None,  # X  30% of Cash Pay - (Pension Contributions)
                        (slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule55').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total +
                         (slip.employee_id.pension
                          and slip.employee_id.pen_contrib or 0.0))
                        if slip.employee_id.resident
                        and slip.employee_id.emp_type in ['primary'] else
                        None,  # Y Actual Pension Contributions including NSSF
                        None,  # Z Permissible Limit (20,000 Ksh)
                        (slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule73').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0)
                        if slip.employee_id.emp_type in ['primary'] else None,
                        # AA
                        (slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule71').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0)
                        if slip.employee_id.emp_type in ['primary'] else None,
                        # AB
                        None,  # AC Amount of Benefit to Deduct before PAYE
                        None,  # AD Taxable Pay
                        None,  # AE Tax Payable
                        (slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule91').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0) if slip.employee_id.resident
                        and slip.employee_id.emp_type in ['primary'] else None,
                        # AF Personal Tax relief
                        (slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule96').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0)
                        if slip.employee_id.resident else None,
                        # AG Insurance Relief
                        None,  # AH PAYE Tax after deduct relief & Insurance Relief - Computed
                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule101').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0  # AI Self Assesed PAYE Tax
                    ]
                    details_employee.append(data_employee)

                # Employees with disability
                if slip.employee_id.disability:
                    data_disabled = [
                        slip.employee_id.tax_pin,  # A KRA PIN for Employee
                        slip.employee_id.name,  # B Employee Full names
                        slip.employee_id.resident and 'Resident'
                        or 'Non-Resident',  # C
                        slip.employee_id.emp_type in ['primary']
                        and 'Primary Employee' or 'Secondary Employee',  # D
                        slip.employee_id.
                        disability_cert,  # E Excemption Cert Number
                        slip.contract_id.wage or 0.0,  # F Basic Salary
                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule17').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # G House Allowance
                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule14').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # H Transport Allowance
                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule15').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,  # I Leave Pay
                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule13').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # J Overtime Allowance
                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule16').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # K Directors Fee
                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule18').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # L Lump Sum Pay
                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule19').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # M Other Allowances
                        None,  # N Total Cash Pay
                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule38').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # O Value of Car benefit
                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule37').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # P sum of Other Benefits(elec,water,telephone,..etc)
                        None,  # Q Total Non Cash Pay
                        (not slip.contract_id.house
                         and str(slip.employee_id.global_income))
                        or (slip.contract_id.house
                            and slip.contract_id.house_type == 'director'
                            and str(slip.employee_id.global_income)) or None,
                        # R Global income (non full time service director)
                        (slip.contract_id.house
                         and slip.contract_id.house_type == 'own'
                         and "Employer's Owned House")
                        or (slip.contract_id.house
                            and slip.contract_id.house_type == 'rented'
                            and "Employer's Rented House")
                        or (slip.contract_id.house
                            and slip.contract_id.house_type == 'agric'
                            and "Agriculture Farm")
                        or (slip.contract_id.house
                            and slip.contract_id.house_type == 'director'
                            and "House to Non full time service Director")
                        or (not slip.contract_id.house
                            and "Benefit not given"),  # S
                        (slip.contract_id.house
                         and slip.contract_id.house_type not in ['director']
                         and slip.contract_id.rent)
                        or (slip.contract_id.house
                            and slip.contract_id.house_type in ['director']
                            and '0.0') or
                        # T Rent of House or Its market Value
                        (not slip.contract_id.house and None),
                        None,  # U Computed Rent of House
                        (slip.contract_id.house
                         and slip.contract_id.house_type not in ['director']
                         and slip.contract_id.rent_recovered)
                        or (slip.contract_id.house
                            and slip.contract_id.house_type in ['director']
                            and '0.0')
                        or (not slip.contract_id.house and None),  # V
                        None,  # W Net value of Housing
                        None,  # X  Total Gross Pay
                        None,  # Y  30% of Cash Pay - (Pension Contributions)
                        (slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule55').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total +
                         (slip.employee_id.pension
                          and slip.employee_id.pen_contrib or 0.0)) or
                        # Z Actual Contributions (NSSF + Pension/Provident)
                        0.0,
                        None,  # AA Permissible Limit (20,000 Ksh)
                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule73').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # AB Mortgage Interest
                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule71').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,  # AC H.O.S.P
                        None,  # AD Amount of Benefit to Deduct before PAYE
                        None,  # AE Exemption for Persons with Disability
                        None,  # AF Taxable Pay
                        None,  # AG Tax Payable
                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule91').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # AH Personal Tax Relief
                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule96').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # AI Insurance Relief
                        None,  # AJ PAYE Tax after deduct personal relief & Insurance Relief
                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              rec.env.ref('hr_ke.ke_rule101').id),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0  # AK Self Assesed PAYE Tax
                    ]
                    details_disabled.append(data_disabled)
            # Write employee details to csv
            if len(details_employee):
                csv_path = self.write_to_csv(details_employee)
                rec.env['hr.ke'].save_attachment(filename_employee, csv_path,
                                                 self._name, rec.id)
            # Write Disabled Employee details to csv
            if len(details_disabled):
                csv_path = self.write_to_csv(details_disabled)
                rec.env['hr.ke'].save_attachment(filename_disabled, csv_path,
                                                 self._name, rec.id)
            # Write car details to csv
            if len(details_cars):
                csv_path = self.write_to_csv(details_cars)
                rec.env['hr.ke'].save_attachment(filename_car, csv_path,
                                                 self._name, rec.id)

    def write_to_csv(self, content):
        csv_fd, csv_path = tempfile.mkstemp(suffix='.csv',
                                            prefix='csvreport.tmp.')
        csv_file = open(csv_path, "w")
        out = csv.writer(csv_file, delimiter=',', quoting=csv.QUOTE_ALL)
        out.writerows(content)
        csv_file.close()
        os.close(csv_fd)
        return csv_path


class KETools(models.Model):
    _name = 'hr.ke'
    _description = "KETools"

    @api.model
    def create_xls(self):
        _, xls_path = tempfile.mkstemp(suffix='.xlsx', prefix='xlsreport.tmp.')
        return xls_path

    @api.model
    def save_attachment(self, filename=None, path=None, res_model=None, res_id=None):
        with open(path, "rb") as report:
            attachment = {
                'name': filename,
                'datas': base64.b64encode(report.read()),
                'res_model': res_model,
                'res_id': res_id,
            }
            try:
                self.env['ir.attachment'].create(attachment)
            except AccessError:
                raise ValidationError(
                    "Cannot save {} as attachment".format(attachment['name']))
            else:
                _logger.info('The document: {} is now saved in the database'.format(
                    attachment['name']))
                self.delete_tempfile(path)

    @api.model
    def delete_tempfile(self, path):
        try:
            os.unlink(path)
        except (OSError, IOError):
            _logger.error('Error when trying to remove file {}'.format(path))

    @api.model
    def style_range(self,
                    ws,
                    cell_range,
                    border=Border(),
                    fill=None,
                    font=None,
                    alignment=None):
        """
        Apply styles to a range of cells as if they were a single cell.

        :param ws:  Excel worksheet instance
        :param range: An excel range to style (e.g. A1:F20)
        :param border: An openpyxl Border
        :param fill: An openpyxl PatternFill or GradientFill
        :param font: An openpyxl Font object
        """

        top = Border(top=border.top)
        left = Border(left=border.left)
        right = Border(right=border.right)
        bottom = Border(bottom=border.bottom)

        first_cell = ws[cell_range.split(":")[0]]
        rows = list(ws[cell_range])
        for cell in rows[0]:
            cell.border = top
        for cell in rows[-1]:
            cell.border = bottom
        for row in rows:
            l = row[0]
            r = row[-1]
            l.border = left
            r.border = right
            if fill:
                for c in row:
                    c.fill = fill
            if font:
                for c in row:
                    c.font = font
            if alignment:
                for c in row:
                    c.alignment = alignment
