import io
import base64
import enum
import logging
import os
import re
import string
import tempfile
from io import BytesIO

from openpyxl.styles import Alignment, Border, Font, PatternFill
from openpyxl.utils import get_column_letter

from odoo import _, api, fields, models
from odoo.exceptions import AccessError, UserError, ValidationError
from odoo.http import request

_logger = logging.getLogger(__name__)
try:
    import openpyxl
except ImportError:
    msg = _(
        'Install python module "openpyxl" in order to create Excel documents')
    raise ValidationError(msg)
try:
    import csv
except ImportError:
    msg = _('Install python module "csv" in order to generate CSV')
    raise ValidationError(msg)

MID_FONT = Font(name='Arial', bold=True, size=12)
NORMAL_FONT = Font(name='Arial', bold=False, size=10)

class HrPayrollFinancialReports(models.Model):
    _inherit = "hr.payslip.run"


    def print_payslip_details(self):

        print('Details')

        """Action to print the Payslip Detailed Report"""
        return self.env.ref('ke_payroll_reports.report_payslip_details').report_action(self)

    def get_nssf_returns(self):
        for rec in self:
            if rec.slip_ids:
                filename_nssf = 'NSSF_Returns-' + re.sub(
                    '[^A-Za-z0-9]+', '',
                    rec.name) + '_' + fields.Datetime.context_timestamp(
                        self, fields.Datetime.now()).strftime(
                            '%Y_%m_%d-%H%M%S') + '.xlsx'
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "NSSF Returns"

                t = 0
                fr = 13

                tiers_m = {}
                tiers_e = {}

                total_income = 0.0
                total_tier1e = 0.0
                total_tier2e = 0.0
                total_tier3e = 0.0
                total_v1e = 0.0

                total_tier1m = 0.0
                total_tier2m = 0.0
                total_tier3m = 0.0
                total_v1m = 0.0

                # EMPLOYER DETAILS
                ws['A1'] = 'RETURNS TYPE'
                # Regular Employees Returns File only, does not include daily
                # paid workers
                ws['B1'] = '01'

                ws['A2'] = 'EMPLOYER KRA PIN'
                # Employer KRA PIN in the HR Settings
                ws['B2'] = self.env.company.vat or None
                
                ws['A3'] = 'EMPLOYER NSSF NUMBER'
                # Employer NSSF Number in the HR/Payroll Settings
                ws['B3'] = self.env.company.nssf or None

                ws['A4'] = 'EMPLOYER NAME'
                ws['B4'] = self.env.company.name or None

                ws['A5'] = 'FISCAL PERIOD'
                # ws['B5'] = datetime.datetime.strptime(
                #    rec.date_end, '%Y-%m-%d').strftime('%m%Y')
                ws['B5'] = fields.Date.from_string(
                    rec.date_end).strftime('%m%Y')


                # DATA HEADERS
                ws['A' + str(fr - 1)] = 'PAYROLL NUMBER'
                ws['B' + str(fr - 1)] = 'SURNAME'
                ws['C' + str(fr - 1)] = 'OTHER NAMES'
                ws['D' + str(fr - 1)] = 'ID NO'
                ws['E' + str(fr - 1)] = 'MEMBER KRA PIN NO'
                ws['F' + str(fr - 1)] = 'NSSF NUMBER'
                ws['G' + str(fr - 1)] = 'CONTRIBUTION TYPE'
                ws['H' + str(fr - 1)] = 'PENSIONABLE INCOME'
                ws['I' + str(fr - 1)] = 'INCOME TYPE'
                ws['J' + str(fr - 1)] = 'MEMBER CONTRIBUTIONS'
                ws['K' + str(fr - 1)] = 'EMPLOYER CONTRIBUTIONS'
                ws['L' + str(fr - 1)] = 'TOTAL CONTRIBUTIONS'

                # DATA ITSELF
                for key, slip in enumerate(rec.slip_ids):
                    if slip.line_ids:
                        income = slip.line_ids.search(
                            [('salary_rule_id.code', '=',
                              'BASIC'),
                             ('slip_id', '=', slip.id)],
                            limit=1).total  # Pensionable income

                    # else:
                    #     msg = self.env._(
                    #         'No Payslip Details!\nPlease compute the payslip for %s'
                    #         % slip.employee_id.name)
                    #     raise ValidationError(msg)


                        tiers_m['t1'] = slip.line_ids.search(
                            [('salary_rule_id.code', '=', 'NSSF_1'),  # Search by salary rule code
                            ('slip_id', '=', slip.id)],
                            limit=1).total  # NSSF Tier I contributions

                        tiers_m['t2'] = slip.line_ids.search(
                            [('salary_rule_id.code', '=', 'NSSF_2'),  # Search by salary rule code
                            ('slip_id', '=', slip.id)],
                            limit=1).total  # NSSF Tier II contributions

                        tiers_m['t3'] = slip.line_ids.search(
                            [('salary_rule_id.code', '=', 'NSSF'),  # Search by salary rule code
                            ('slip_id', '=', slip.id)],
                            limit=1).total  # NSSF Total contributions


                        # tiers_m['v1'] = slip.line_ids.search(
                        #     [('salary_rule_id', '=',
                        #       rec.env.ref('hr_ke.ke_rule49').id),
                        #      ('slip_id', '=', slip.id)],
                        #     limit=1).total  # NSSF Voluntary contributions


                        # Employer Contributions
                        tiers_e['t1'] = slip.line_ids.search(
                            [('salary_rule_id.code', '=', 'NSSF_1'),  # Search by salary rule code
                            ('slip_id', '=', slip.id)],
                            limit=1).total  # NSSF Tier I contributions


                        tiers_e['t2'] = slip.line_ids.search(
                            [('salary_rule_id.code', '=', 'NSSF_2'),  # Search by salary rule code
                            ('slip_id', '=', slip.id)],
                            limit=1).total  # NSSF Tier II contributions

                        tiers_e['t3'] = slip.line_ids.search(
                            [('salary_rule_id.code', '=', 'NSSF'),  # Search by salary rule code
                            ('slip_id', '=', slip.id)],
                            limit=1).total  # NSSF Total contributions


                        # tiers_e['v1'] = slip.line_ids.search(
                        #     [('salary_rule_id', '=',
                        #       rec.env.ref('hr_ke.ke_rule59').id),
                        #      ('slip_id', '=', slip.id)],
                        #     limit=1).total  # NSSF Voluntary contributions


                        # Dermine the TOTALS
                        total_income += income
                        total_tier1e += tiers_e['t1']
                        total_tier2e += tiers_e['t2']
                        total_tier3e += tiers_e['t3']
                        # total_v1e += tiers_e['v1']

                        total_tier1m += tiers_m['t1']
                        total_tier2m += tiers_m['t2']
                        total_tier3m += tiers_m['t3']
                        # total_v1m += tiers_m['v1']

                        tiers_m = {
                            x: tiers_m[x]
                            for x in tiers_m.keys() if (tiers_m[x] or tiers_e[x])
                        }  # Ignore Zero Contributions from Member
                        # Ignore Zero Contributions from Employer
                        tiers_e = {
                            x: tiers_e[x]
                            for x in tiers_m.keys() if (tiers_e[x] or tiers_m[x])
                        }
                        for k in sorted(tiers_m):
                            # PAYROLL NUMBER
                            ws['A' + str(fr + key +
                                        t)] = slip.employee_id.employee_no or None
                            # SURNAME
                            ws['B' + str(fr + key +
                                        t)] = slip.employee_id.display_name.split(
                                            ' ')[-1] or ''
                            # OTHER NAMES
                            ws['C' + str(fr + key +
                                        t)] = slip.employee_id.display_name.split(
                                            ' ')[0] or ''
                            # ID NO (National id number/Alien registration
                            # no./Passport no.)
                            ws['D' + str(
                                fr + key + t
                            )] = slip.employee_id.identification_id or slip.employee_id.passport_id or None
                            # MEMBER KRA PIN NO
                            ws['E' + str(fr + key +
                                        t)] = slip.employee_id.kra_pin or None
                            # NSSF NUMBER
                            ws['F' +
                            str(fr + key + t)] = slip.employee_id.nssf or None
                            # CONTRIBUTION TYPE (Value range: 101, 102, 103, 104,
                            # 105, 200)
                            ws['G' + str(
                                fr + key + t
                            )] = k == 't1' and 101 or k == 't2' and 102 or k == 't3' and 103 or k == 'v1' and 105 or None
                            # PENSIONABLE INCOME {AS DEFINED BY THE NSSF ACT 20fr }
                            ws['H' +
                            str(fr + key + t)] = k in ['t1', 't2', 't3'
                                                        ] and income or None
                            # INCOME TYPE (Value range: 1, 2, 3, 4) ....'1' becuase
                            # this return is for for monthly paid workers only
                            ws['I' + str(fr + key + t)] = 1
                            # MEMBER CONTRIBUTIONS
                            ws['J' + str(fr + key + t)] = tiers_m[k]
                            # EMPLOYER CONTRIBUTIONS
                            ws['K' + str(fr + key + t)] = tiers_e[k]
                            ws['L' +
                            str(fr + key +
                                t)] = ws['J' + str(fr + key + t)].value + ws[
                                    'K' + str(fr + key + t)].value
                            t += 1
                        t -= 1

                # SUMMARY DETAILS
                ws['A6'] = 'TOTAL SALARIES'
                ws['B6'] = total_income
                ws['A7'] = 'TOTAL MEMBER CONTRIBUTIONS'
                ws['B7'] = (total_tier1m + total_tier2m + total_tier3m +
                            total_v1m)
                ws['A8'] = 'TOTAL EMPLOYER CONTRIBUTIONS'
                ws['B8'] = (total_tier1e + total_tier2e + total_tier3e +
                            total_v1e)
                ws['A9'] = 'TOTAL CONTRIBUTIONS'
                ws['B9'] = ws['B7'].value + ws['B8'].value
                ws['A10'] = 'NO OF RECORDS'
                ws['B10'] = t + len(rec.slip_ids)
                
                # Ensure the directory exists
                xls_path = "/tmp/xlsreport.xlsx"


                if not os.path.exists(xls_path):
                    # Create an empty file
                    with open(xls_path, 'wb') as temp_file:
                        temp_file.write(b'')

                # Open the file and write the Excel data
                with open(xls_path, 'wb') as f:
                    wb.save(f)  # Save workbook to file

                # Read the generated file from the path
                with open(xls_path, 'rb') as f:
                    file_data = base64.b64encode(f.read())

                # Create an attachment
                attachment = self.env['ir.attachment'].create({
                    'name': filename_nssf,
                    'type': 'binary',
                    'datas': file_data,
                    'res_model': 'hr.payslip.run',  # Replace with your model name
                    'res_id': rec.id,  # Attach to the current record
                })

                # Return the file as a downloadable attachment
                return {
                    'type': 'ir.actions.act_url',
                    'url': '/web/content/%s?download=true' % attachment.id,
                    'target': 'self',
                }
            else:
                msg = _('No Payslips to process!')
                raise ValidationError(msg)

    
    def get_shif_returns(self):
        for rec in self:
            if rec.slip_ids:
                filename_nhif = 'SHIF_Returns-' + re.sub(
                    '[^A-Za-z0-9]+', '',
                    rec.name) + '_' + fields.Datetime.context_timestamp(
                        self, fields.Datetime.now()).strftime(
                            '%Y_%m_%d-%H%M%S') + '.xlsx'

                wb = openpyxl.Workbook()
                ws = wb.active

                fr = 6
                total = 0.0
                nhif = 0.0

                # EMPLOYER DETAILS
                ws['A1'] = 'EMPLOYER CODE'
                # NHIF NO IN THE HR SETTINGS
                ws['B1'] = self.env.company.nhif or None

                ws['A2'] = 'EMPLOYER NAME'
                ws['B2'] = self.env.company.name or None

                ws['A3'] = 'MONTH OF CONTRIBUTION'
                # ws['B3'] = datetime.datetime.strptime(
                #    rec.date_end, '%Y-%m-%d').strftime('%Y-%m')
                ws['B3'] = fields.Date.from_string(
                    rec.date_end).strftime('%Y-%m')

                # DATA HEADERS
                ws['A' + str(fr - 1)] = 'PAYROLL NO'
                ws['B' + str(fr - 1)] = 'LAST NAME'
                ws['C' + str(fr - 1)] = 'FIRST NAME'
                ws['D' + str(fr - 1)] = 'ID NO'
                ws['E' + str(fr - 1)] = 'SHIF NO'
                ws['F' + str(fr - 1)] = 'AMOUNT'

                # DATA ITSELF
                for key, slip in enumerate(rec.slip_ids):
                    if slip.line_ids:
                        nhif = slip.line_ids.search(
                            [('salary_rule_id.code', '=',
                              'SHIF'),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0  # SHIF contributions
                    else:
                        msg = _(
                            'No Payslip Details!\nPlease compute the payslip for %s'
                            % slip.employee_id.name)
                        raise ValidationError(msg)
                        
                    # CELLS
                    # PAYROLL NUMBER
                    name = slip.employee_id.name.strip().split(' ')
                    ws['A' +
                       str(fr + key)] = slip.employee_id.employee_no or None
                    # LAST NAME
                    ws['B' +
                       str(fr + key)] = name[-1] or ''
                    # FIRST NAME
                    ws['C' +
                       str(fr + key)] = name[0] or ''
                    # ID NO (National id number/Alien registration no./Passport
                    # no.)
                    ws['D' + str(
                        fr + key
                    )] = slip.employee_id.identification_id or slip.employee_id.passport_id or None
                    # NHIF Number
                    ws['E' + str(fr + key)] = slip.employee_id.nhif or None
                    ws['F' + str(fr + key)] = nhif  # Amount contributed
                    total += nhif
                # TOTAL
                ws['E' + str(fr + key + 1)] = 'TOTAL'
                ws['F' + str(fr + key + 1)] = total


                # Ensure the directory exists
                xls_path = "/tmp/xlsreport.xlsx"


                if not os.path.exists(xls_path):
                    # Create an empty file
                    with open(xls_path, 'wb') as temp_file:
                        temp_file.write(b'')

                # Open the file and write the Excel data
                with open(xls_path, 'wb') as f:
                    wb.save(f)  # Save workbook to file

                # Read the generated file from the path
                with open(xls_path, 'rb') as f:
                    file_data = base64.b64encode(f.read())

                # Create an attachment
                attachment = self.env['ir.attachment'].create({
                    'name': filename_nhif,
                    'type': 'binary',
                    'datas': file_data,
                    'res_model': 'hr.payslip.run',
                    'res_id': rec.id,  # Attach to the current record
                })

                # Return the file as a downloadable attachment
                return {
                    'type': 'ir.actions.act_url',
                    'url': '/web/content/%s?download=true' % attachment.id,
                    'target': 'self',
                }
            else:
                msg = _('No Payslips to process!')
                raise ValidationError(msg)



    def _prepare_payroll_summary_data(self):
        data = []
        rules = [
            ('BASIC PAY', 'BASIC'), ('HOUSING LEVY', 'HOUSING'), ('SHIF', 'SHIF'), ('NSSF', 'NSSF'), ('TAXABLE PAY', 'TAXABLEPAY'),
            ('INCOME TAX', 'TAXES'), ('PERSONAL RELIEF', 'RELIEF'), ('PAYE', 'PAYE'), ('NET PAY', 'NET')
            # ('NOTICE PAY', 'SA15'),
            # ('REIMBURSEMENT', 'P016'), ('BONUS', 'P011'), ('ABSENT (Units)', 'P0900'),
            # ('ABSENT (Amount)', 'PI01'), ('GROSS PAY', 'P030'), ('PAYE', 'P101'), ('NSSF', 'P055'), ('NHIF', 'P106'), ('NHDL', 'P111'),
            # ('HELB', 'P107'), ('SALARY ADVANCE', 'P108'), ('COMPANY LOAN', 'LOANINS'), ('SACCO', 'P109'), ('SURCHARGE/RECOVERY', 'P113'),
            # ('NOTICE DEDUCTION', 'P114'), ('TOTAL DEDUCTIONS', 'P115'), ('NET PAY', 'P120')
        ]

        for index, slip in enumerate(self.slip_ids, start=1):
            employee = slip.employee_id
            vals = {
                'SR.': index,
                'PAYROLL NUMBER': employee.employee_no or None,
                'EMPLOYEE NAME': employee.name,
                'BANK NAME': employee.bank_branch or None,
                'BANK BRANCH CODE': employee.bank_code or None,
                'BANK ACCOUNT NUMBER': employee.account_number or None
            }
            for rule, code in rules:
                vals[rule] = slip.line_ids.filtered(lambda line: line.code == code).total
            data.append(vals)
        return data
    
    def get_payroll_summary(self):
        for rec in self:
            filename = f'{self.name} Payroll_Summary-.xlsx'

            wb = openpyxl.Workbook()
            ws = wb.active

            fr = 1
            employee_cols = ['SR.', 'PAYROLL NUMBER', 'EMPLOYEE NAME']

            rules_cols = [
                'BASIC PAY', 'HOUSING LEVY', 'SHIF', 'NSSF', 'TAXABLE PAY',
                'INCOME TAX', 'PERSONAL RELIEF', 'PAYE', 'NET PAY'
                # 'NOTICE PAY', 'REIMBURSEMENT', 'BONUS', 'ABSENT (Units)',
                # 'ABSENT (Amount)', 'GROSS PAY', 'PAYE', 'NSSF', 'NHIF', 'NHDL', 'HELB', 'SALARY ADVANCE', 'COMPANY LOAN', 'SACCO', 'SURCHARGE/RECOVERY',
                # 'NOTICE DEDUCTION',	'TOTAL DEDUCTIONS',	'NET PAY'
            ]

            bank_cols =['BANK NAME', 'BANK BRANCH CODE', 'BANK ACCOUNT NUMBER']

            cols = employee_cols + rules_cols + bank_cols
            
            for index, col in enumerate(cols, start=1):
                letter = get_column_letter(index)
                ws[f'{letter}{fr}'] = col
                ws[f'{letter}{fr}'].font = MID_FONT
            fr += 1
            start_fr = fr
            for sumr in self._prepare_payroll_summary_data():
                for index, col in enumerate(cols, start=1):
                    letter = get_column_letter(index)
                    ws[f'{letter}{fr}'] = sumr[col]
                    ws[f'{letter}{fr}'].font = NORMAL_FONT
                fr += 1
            end_fr = fr - 1
            
            employee_cols_end = len(employee_cols)
            employee_cols_letter = get_column_letter(employee_cols_end)
            ws[f'{employee_cols_letter}{fr}'] = 'TOTALS'
            
            for index, col in enumerate(rules_cols, start = employee_cols_end + 1):
                letter = get_column_letter(index)
                ws[f'{letter}{fr}'] = f'=SUM({letter}{start_fr}:{letter}{end_fr})'
                ws[f'{letter}{fr}'].font = MID_FONT
            

                # Ensure the directory exists
                xls_path = "/tmp/xlsreport.xlsx"


                if not os.path.exists(xls_path):
                    # Create an empty file
                    with open(xls_path, 'wb') as temp_file:
                        temp_file.write(b'')

                # Open the file and write the Excel data
                with open(xls_path, 'wb') as f:
                    wb.save(f)  # Save workbook to file

                # Read the generated file from the path
                with open(xls_path, 'rb') as f:
                    file_data = base64.b64encode(f.read())

                # Create an attachment
                attachment = self.env['ir.attachment'].create({
                    'name': filename,
                    'type': 'binary',
                    'datas': file_data,
                    'res_model': 'hr.payslip.run',
                    'res_id': rec.id,  # Attach to the current record
                })

                # Return the file as a downloadable attachment
                return {
                    'type': 'ir.actions.act_url',
                    'url': '/web/content/%s?download=true' % attachment.id,
                    'target': 'self',
                }


    def get_net_pay(self):
        for rec in self:
            if rec.slip_ids:
                filename_netpay = 'NET_PAY-' + re.sub(
                    '[^A-Za-z0-9]+', '',
                    rec.name) + '_' + fields.Datetime.context_timestamp(
                        self, fields.Datetime.now()).strftime(
                            '%Y_%m_%d-%H%M%S') + '.xlsx'

                wb = openpyxl.Workbook()
                ws = wb.active

                fr = 7  # First row of data

                ws['A1'] = self.env.company.name
                ws['A2'] = 'PAYROLL SUMMARY'
                ws['B2'] = rec.name
                
                cols = [
                    'EMPLOYEE NAME', 'ACCOUNT NO', 'BANK BRANCH',
                    'AMOUNT'
                ]


                # DATA HEADERS
                for k, x in enumerate(string.ascii_uppercase[0:4]):  # 'ABCD'
                    ws[x + str(fr - 1)] = cols[k]

                for key, slip in enumerate(rec.slip_ids):
                    ws['A' + str(fr + key)] = slip.employee_id.name or None
                    ws['B' + str(
                        fr + key
                    )] = slip.employee_id.account_number or None
                    ws['C' + str(
                        fr + key
                    )] = slip.employee_id.bank_branch or None
                    # ws['D' + str(
                    #     fr + key
                    # )] = slip.employee_id.bank_account_id.bank_id.bic or None

                    ws['D' + str(fr + key)] = slip.line_ids.search(
                        [('salary_rule_id.code', '=',
                          'NET'),
                         ('slip_id', '=', slip.id)],
                        limit=1).total  # Total Net Pay
                        
                # Totals
                t = fr + key + 1  # last row for Totals

                ws['C' + str(t)] = 'TOTAL'
                # Sum using excel 'SUM' formula
                ws['D' + str(t)] = '=SUM(D' + str(fr) + ':D' + str(t - 1) + ')'

                # Ensure the directory exists
                xls_path = "/tmp/xlsreport.xlsx"


                if not os.path.exists(xls_path):
                    # Create an empty file
                    with open(xls_path, 'wb') as temp_file:
                        temp_file.write(b'')

                # Open the file and write the Excel data
                with open(xls_path, 'wb') as f:
                    wb.save(f)  # Save workbook to file

                # Read the generated file from the path
                with open(xls_path, 'rb') as f:
                    file_data = base64.b64encode(f.read())

                # Create an attachment
                attachment = self.env['ir.attachment'].create({
                    'name': filename_netpay,
                    'type': 'binary',
                    'datas': file_data,
                    'res_model': 'hr.payslip.run',
                    'res_id': rec.id,  # Attach to the current record
                })

                # Return the file as a downloadable attachment
                return {
                    'type': 'ir.actions.act_url',
                    'url': '/web/content/%s?download=true' % attachment.id,
                    'target': 'self',
                }
            else:
                msg = _('No Payslips to process!')
                raise ValidationError(msg)


    def get_p10(self):
        for rec in self:
            filename_employee = 'Employees_Details-' + re.sub(
                '[^A-Za-z0-9]+', '',
                rec.name) + '_' + fields.Datetime.context_timestamp(
                    self,
                    fields.Datetime.now()).strftime('%Y_%m_%d-%H%M%S') + '.csv'
            filename_disabled = 'Disabled_Employees_Details-' + re.sub(
                '[^A-Za-z0-9]+', '',
                rec.name) + '_' + fields.Datetime.context_timestamp(
                    self,
                    fields.Datetime.now()).strftime('%Y_%m_%d-%H%M%S') + '.csv'
            filename_car = 'Car_Benefit_Details-' + re.sub(
                '[^A-Za-z0-9]+', '',
                rec.name) + '_' + fields.Datetime.context_timestamp(
                    self,
                    fields.Datetime.now()).strftime('%Y_%m_%d-%H%M%S') + '.csv'
            details_employee = []
            details_disabled = []
            details_cars = []
            for slip in rec.slip_ids:
                if slip.contract_id.car and slip.contract_id.cars:
                    for car in slip.contract_id.cars:
                        data_cars = [
                            slip.employee_id.kra_pin,  # A
                            slip.employee_id.disability
                            and 'C_Disabled_Employees_Dtls'
                            or 'B_Employees_Dtls',  # B
                            car.name,  # C
                            car.make,  # D
                            car.body in ['saloon']
                            and 'Saloon Hatch Backs and Estates'
                            or car.body in ['pickup']
                            and 'Pick Ups, Panel Vans Uncovered'
                            or car.body in ['cruiser'] and
                            # E
                            'Land Rovers/ Cruisers(excludes Range Rovers and vehicles of similar nature)',
                            car.cc_rate,  # F
                            car.cost_type,  # G
                            car.cost_type in ['Hired'] and car.cost_hire
                            or None,  # H
                            car.cost_type in ['Owned'] and car.cost_own
                            or None,  # I
                        ]
                        details_cars.append(data_cars)

                # Normal Employees without disability
                if not slip.employee_id.disability:
                    if not slip.employee_id.kra_pin:
                        raise ValidationError(
                            _('KRA PIN Number for %s is missing!' %
                              slip.employee_id.name))

                    data_employee = [
                        slip.employee_id.kra_pin,  # A
                        slip.employee_id.name,  # B
                        slip.employee_id.resident and 'Resident'
                        or 'Non-Resident',  # C
                        slip.employee_id.emp_type in ['primary']
                        and 'Primary Employee' or 'Secondary Employee',  # D
                        slip.contract_id.wage or 0.0,  # E
                        slip.line_ids.search(
                            [('salary_rule_id.code', '=',
                              'HA'), # house allowance
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # F House Allowance

                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              'IA'), # inconvenience allowance
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # G Transport Allowance

                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              'LP'), # leave pay
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,  # H Leave Pay

                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              'OTA'), # overtime allowance
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # I Overtime Allowance

                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              'RA'), # Reimbursement Alowance
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # J Directors Fee

                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              'LSP'), # Lump Sum Pay
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # K Lump Sum Pay

                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              'OA'), # Other Allowances
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # L Other Allowances
                        None,
                        
                        # M
                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              'CB'), # Car Benefit
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # N Value of Car Benefit

                        slip.line_ids.search(
                            [('salary_rule_id', '=',
                              'NCB'), # Non cash benefits
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # O sum of other benefits (water,elec,telephone..etc)
                        None,  # P Total non cash pay -computed by P10 Form
                        (not slip.contract_id.house
                         and str(slip.employee_id.global_income))
                        or (slip.contract_id.house
                            and slip.contract_id.house_type == 'director'
                            and str(slip.employee_id.global_income)) or None,
                        # Q Global income (non full time service director)
                        (slip.contract_id.house
                         and slip.contract_id.house_type == 'own'
                         and "Employer's Owned House")
                        or (slip.contract_id.house
                            and slip.contract_id.house_type == 'rented'
                            and "Employer's Rented House")
                        or (slip.contract_id.house
                            and slip.contract_id.house_type == 'agric'
                            and "Agriculture Farm")
                        or (slip.contract_id.house
                            and slip.contract_id.house_type == 'director'
                            and "House to Non full time service Director")
                        or (not slip.contract_id.house
                            and "Benefit not given"),  # R
                        (slip.contract_id.house
                         and slip.contract_id.house_type not in ['director']
                         and slip.contract_id.rent)
                        or (slip.contract_id.house
                            and slip.contract_id.house_type in ['director']
                            and '0.0') or
                        # S Rent of House or Its market Value
                        (not slip.contract_id.house and None),
                        None,  # T Computed Rent of house(15% of GrossPay)
                        (slip.contract_id.house
                         and slip.contract_id.house_type not in ['director']
                         and slip.contract_id.rent_recovered)
                        or (slip.contract_id.house
                            and slip.contract_id.house_type in ['director']
                            and '0.0') or
                        # U Rent Recovered from Employee
                        (not slip.contract_id.house and None),
                        None,  # V Net value of Housing
                        None,  # W  Total Gross Pay
                        None,  # X  30% of Cash Pay - (Pension Contributions)
                        (slip.line_ids.search(
                            [('salary_rule_id.code', '=',
                              'NSSF'),
                             ('slip_id', '=', slip.id)],
                            limit=1).total +
                         (slip.employee_id.pension
                          and slip.employee_id.pen_contrib or 0.0))
                        if slip.employee_id.resident
                        and slip.employee_id.emp_type in ['primary'] else
                        None,  # Y Actual Pension Contributions including NSSF
                        None,  # Z Permissible Limit (20,000 Ksh)
                        (slip.line_ids.search(
                            [('salary_rule_id.code', '=',
                              'MI'), # Mortgage Interest
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0)

                        if slip.employee_id.emp_type in ['primary'] else None,
                        # AA
                        (slip.line_ids.search(
                            [('salary_rule_id.code', '=',
                              'HOSP'), #HOSP
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0)
                        if slip.employee_id.emp_type in ['primary'] else None,
                        # AB
                        None,  # AC Amount of Benefit to Deduct before PAYE
                        None,  # AD Taxable Pay
                        None,  # AE Tax Payable
                        (slip.line_ids.search(
                            [('salary_rule_id.code', '=',
                              'RELIEF'), # Personal Tax Relief
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0) if slip.employee_id.resident
                        and slip.employee_id.emp_type in ['primary'] else None,
                        # AF Personal Tax relief
                        (slip.line_ids.search(
                            [('salary_rule_id.code', '=',
                              'IR'), # Insurance Relief
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0)
                        if slip.employee_id.resident else None,
                        # AG Insurance Relief
                        None,  # AH PAYE Tax after deduct relief & Insurance Relief - Computed
                        slip.line_ids.search(
                            [('salary_rule_id.code', '=',
                              'PAYE'),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0  # AI Self Assesed PAYE Tax
                    ]
                    details_employee.append(data_employee)

                # Employees with disability
                if slip.employee_id.disability:
                    data_disabled = [
                        slip.employee_id.kra_pin,  # A KRA PIN for Employee
                        slip.employee_id.name,  # B Employee Full names
                        slip.employee_id.resident and 'Resident'
                        or 'Non-Resident',  # C
                        slip.employee_id.emp_type in ['primary']
                        and 'Primary Employee' or 'Secondary Employee',  # D
                        slip.employee_id.
                        disability_cert,  # E Excemption Cert Number
                        slip.contract_id.wage or 0.0,  # F Basic Salary
                        slip.line_ids.search(
                            [('salary_rule_id.code', '=',
                              'HA'), # House Allowance
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # G House Allowance

                        slip.line_ids.search(
                            [('salary_rule_id.code', '=',
                              'IA'), # Inconvenience Allowance / Transport Allowance
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # H Transport Allowance

                        slip.line_ids.search(
                            [('salary_rule_id.code', '=',
                              'LP'), # Leave Pay
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,  # I Leave Pay

                        slip.line_ids.search(
                            [('salary_rule_id.code', '=',
                              'OTA'), # Over Time Allowance
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # J Overtime Allowance

                        slip.line_ids.search(
                            [('salary_rule_id.code', '=',
                            'RA'), # Reimbursement Allowance
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # K Directors Fee

                        slip.line_ids.search(
                            [('salary_rule_id.code', '=',
                              'LSP'), # Lump sum pay
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # L Lump Sum Pay

                        slip.line_ids.search(
                            [('salary_rule_id.code', '=',
                              'OA'), # Other Allowances
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # M Other Allowances
                        None,  # N Total Cash Pay

                        slip.line_ids.search(
                            [('salary_rule_id.code', '=',
                              'MVA'), # Motor Vehicle Allowance
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # O Value of Car benefit

                        slip.line_ids.search(
                            [('salary_rule_id.code', '=',
                              'NCB'), # Non cash benefits
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # P sum of Other Benefits(elec,water,telephone,..etc)
                        None,  # Q Total Non Cash Pay

                        (not slip.contract_id.house
                         and str(slip.employee_id.global_income))
                        or (slip.contract_id.house
                            and slip.contract_id.house_type == 'director'
                            and str(slip.employee_id.global_income)) or None,
                        # R Global income (non full time service director)
                        (slip.contract_id.house
                         and slip.contract_id.house_type == 'own'
                         and "Employer's Owned House")
                        or (slip.contract_id.house
                            and slip.contract_id.house_type == 'rented'
                            and "Employer's Rented House")
                        or (slip.contract_id.house
                            and slip.contract_id.house_type == 'agric'
                            and "Agriculture Farm")
                        or (slip.contract_id.house
                            and slip.contract_id.house_type == 'director'
                            and "House to Non full time service Director")
                        or (not slip.contract_id.house
                            and "Benefit not given"),  # S
                        (slip.contract_id.house
                         and slip.contract_id.house_type not in ['director']
                         and slip.contract_id.rent)
                        or (slip.contract_id.house
                            and slip.contract_id.house_type in ['director']
                            and '0.0') or
                        # T Rent of House or Its market Value
                        (not slip.contract_id.house and None),
                        None,  # U Computed Rent of House
                        (slip.contract_id.house
                         and slip.contract_id.house_type not in ['director']
                         and slip.contract_id.rent_recovered)
                        or (slip.contract_id.house
                            and slip.contract_id.house_type in ['director']
                            and '0.0')
                        or (not slip.contract_id.house and None),  # V
                        None,  # W Net value of Housing
                        None,  # X  Total Gross Pay
                        None,  # Y  30% of Cash Pay - (Pension Contributions)
                        (slip.line_ids.search(
                            [('salary_rule_id.code', '=',
                              'NSSF'),
                             ('slip_id', '=', slip.id)],
                            limit=1).total +
                         (slip.employee_id.pension
                          and slip.employee_id.pen_contrib or 0.0)) or
                        # Z Actual Contributions (NSSF + Pension/Provident)
                        0.0,
                        None,  # AA Permissible Limit (20,000 Ksh)

                        slip.line_ids.search(
                            [('salary_rule_id.code', '=',
                              'MI'), # Mortgage interest
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # AB Mortgage Interest

                        slip.line_ids.search(
                            [('salary_rule_id.code', '=',
                              'HOSP'),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,  # AC H.O.S.P
                        None,  # AD Amount of Benefit to Deduct before PAYE
                        None,  # AE Exemption for Persons with Disability
                        None,  # AF Taxable Pay
                        None,  # AG Tax Payable

                        slip.line_ids.search(
                            [('salary_rule_id.code', '=',
                              'RELIEF'),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # AH Personal Tax Relief

                        slip.line_ids.search(
                            [('salary_rule_id.code', '=',
                              'IR'), # Insurance Relief
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0,
                        # AI Insurance Relief
                        None,  # AJ PAYE Tax after deduct personal relief & Insurance Relief

                        slip.line_ids.search(
                            [('salary_rule_id.code', '=',
                              'PAYE'),
                             ('slip_id', '=', slip.id)],
                            limit=1).total or 0.0  # AK Self Assesed PAYE Tax
                    ]
                    details_disabled.append(data_disabled)
            # Write employee details to csv
            if len(details_employee):
                csv_path = self.write_to_csv(details_employee)
                rec.env['hr.ke'].save_attachment(filename_employee, csv_path,
                                                 self._name, rec.id)
                

                # Ensure the directory exists
                csv_path = "/tmp/xlsreport.xlsx"


                if not os.path.exists(csv_path):
                    # Create an empty file
                    with open(csv_path, 'wb') as temp_file:
                        temp_file.write(b'')

                attachment = self.env['ir.attachment'].create({
                    'name': filename_employee,
                    'type': 'binary',
                    'datas': base64.b64encode(open(csv_path, 'rb').read()),
                    'res_model': self._name,
                    'res_id': self.id,
                })
                attachment_id = attachment.id

            
            # Write Disabled Employee details to csv
            if len(details_disabled):
                csv_path = self.write_to_csv(details_disabled)
                rec.env['hr.ke'].save_attachment(filename_disabled, csv_path,
                                                 self._name, rec.id)

                # Ensure the directory exists
                csv_path = "/tmp/xlsreport.xlsx"


                if not os.path.exists(csv_path):
                    # Create an empty file
                    with open(csv_path, 'wb') as temp_file:
                        temp_file.write(b'')

                attachment = self.env['ir.attachment'].create({
                    'name': filename_disabled,
                    'type': 'binary',
                    'datas': base64.b64encode(open(csv_path, 'rb').read()),
                    'res_model': self._name,
                    'res_id': self.id,
                })
                attachment_id = attachment.id

            # Write car details to csv
            if len(details_cars):
                csv_path = self.write_to_csv(details_cars)
                rec.env['hr.ke'].save_attachment(filename_car, csv_path,
                                                 self._name, rec.id)

                # Ensure the directory exists
                csv_path = "/tmp/xlsreport.xlsx"


                if not os.path.exists(csv_path):
                    # Create an empty file
                    with open(csv_path, 'wb') as temp_file:
                        temp_file.write(b'')
                
                attachment = self.env['ir.attachment'].create({
                    'name': filename_car,
                    'type': 'binary',
                    'datas': base64.b64encode(open(csv_path, 'rb').read()),
                    'res_model': self._name,
                    'res_id': self.id,
                })
                attachment_id = attachment.id

            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/{attachment_id}?download=true',
                'target': 'self',
            } if attachment_id else None

    def write_to_csv(self, content):
        csv_fd, csv_path = tempfile.mkstemp(suffix='.csv',
                                            prefix='csvreport.tmp.')
        csv_file = open(csv_path, "w")
        out = csv.writer(csv_file, delimiter=',', quoting=csv.QUOTE_ALL)
        out.writerows(content)
        csv_file.close()
        os.close(csv_fd)
        return csv_path


    def get_helb(self):
        for rec in self:
            if rec.slip_ids:
                filename_netpay = 'HELB-' + re.sub(
                    '[^A-Za-z0-9]+', '',
                    rec.name) + '_' + fields.Datetime.context_timestamp(
                        self, fields.Datetime.now()).strftime(
                            '%Y_%m_%d-%H%M%S') + '.xlsx'

                wb = openpyxl.Workbook()
                ws = wb.active

                fr = 7  # First row of data

                ws['A1'] = self.env.company.name
                ws['A2'] = 'HELB REPORT'
                ws['B2'] = rec.name
                
                cols = [
                    'ID NUMBER', 'EMPLOYEE NAME', 'STAFF NUMBER',
                    'HELB AMOUNT'
                ]


                # DATA HEADERS
                for k, x in enumerate(string.ascii_uppercase[0:4]):  # 'ABCD'
                    ws[x + str(fr - 1)] = cols[k]

                for key, slip in enumerate(rec.slip_ids):
                    ws['A' + str(fr + key)] = slip.employee_id.identification_id or None

                    ws['B' + str(fr + key)] = slip.employee_id.name or None

                    ws['C' + str(
                        fr + key
                    )] = slip.employee_id.employee_no or None

                    # ws['D' + str(
                    #     fr + key
                    # )] = slip.employee_id.helb_rate or None

                    # ws['D' + str(
                    #     fr + key
                    # )] = slip.employee_id.bank_account_id.bank_id.bic or None

                    helb = slip.line_ids.search(
                        [('salary_rule_id.code', '=',
                          'HELB'),
                         ('slip_id', '=', slip.id)],
                        limit=1).total  # HELB

                    print("HELB 2: ", helb)

                    ws['D' + str(fr + key)] = helb
                        
                # Totals
                t = fr + key + 1  # last row for Totals

                ws['C' + str(t)] = 'TOTAL'
                # Sum using excel 'SUM' formula
                ws['D' + str(t)] = '=SUM(D' + str(fr) + ':D' + str(t - 1) + ')'

                # Ensure the directory exists
                xls_path = "/tmp/xlsreport.xlsx"


                if not os.path.exists(xls_path):
                    # Create an empty file
                    with open(xls_path, 'wb') as temp_file:
                        temp_file.write(b'')

                # Open the file and write the Excel data
                with open(xls_path, 'wb') as f:
                    wb.save(f)  # Save workbook to file

                # Read the generated file from the path
                with open(xls_path, 'rb') as f:
                    file_data = base64.b64encode(f.read())

                # Create an attachment
                attachment = self.env['ir.attachment'].create({
                    'name': filename_netpay,
                    'type': 'binary',
                    'datas': file_data,
                    'res_model': 'hr.payslip.run',
                    'res_id': rec.id,  # Attach to the current record
                })

                # Return the file as a downloadable attachment
                return {
                    'type': 'ir.actions.act_url',
                    'url': '/web/content/%s?download=true' % attachment.id,
                    'target': 'self',
                }
            else:
                msg = _('No Payslips to process!')
                raise ValidationError(msg)


    def get_paye(self):
            for rec in self:
                if rec.slip_ids:
                    filename_netpay = 'PAYE-' + re.sub(
                        '[^A-Za-z0-9]+', '',
                        rec.name) + '_' + fields.Datetime.context_timestamp(
                            self, fields.Datetime.now()).strftime(
                                '%Y_%m_%d-%H%M%S') + '.xlsx'

                    wb = openpyxl.Workbook()
                    ws = wb.active

                    fr = 7  # First row of data

                    ws['A1'] = self.env.company.name
                    ws['A2'] = 'PAYE REPORT'
                    ws['A3'] = 'Month'
                    ws['B3'] = rec.name
                    
                    cols = [
                        'ID NUMBER', 'EMPLOYEE NAME', 'KRA PIN',
                        'PAYE AMOUNT'
                    ]


                    # DATA HEADERS
                    for k, x in enumerate(string.ascii_uppercase[0:4]):  # 'ABCD'
                        ws[x + str(fr - 1)] = cols[k]

                    for key, slip in enumerate(rec.slip_ids):
                        ws['A' + str(fr + key)] = slip.employee_id.identification_id or None

                        ws['B' + str(fr + key)] = slip.employee_id.name or None

                        ws['C' + str(
                            fr + key
                        )] = slip.employee_id.kra_pin or None


                        helb = slip.line_ids.search(
                            [('salary_rule_id.code', '=',
                            'PAYE'),
                            ('slip_id', '=', slip.id)],
                            limit=1).total  # PAYE

                        ws['D' + str(fr + key)] = helb
                            
                    # Totals
                    t = fr + key + 2  # last row for Totals

                    ws['C' + str(t)] = 'TOTAL'
                    # Sum using excel 'SUM' formula
                    ws['D' + str(t)] = '=SUM(D' + str(fr) + ':D' + str(t - 1) + ')'

                    # Ensure the directory exists
                    xls_path = "/tmp/xlsreport.xlsx"


                    if not os.path.exists(xls_path):
                        # Create an empty file
                        with open(xls_path, 'wb') as temp_file:
                            temp_file.write(b'')

                    # Open the file and write the Excel data
                    with open(xls_path, 'wb') as f:
                        wb.save(f)  # Save workbook to file

                    # Read the generated file from the path
                    with open(xls_path, 'rb') as f:
                        file_data = base64.b64encode(f.read())

                    # Create an attachment
                    attachment = self.env['ir.attachment'].create({
                        'name': filename_netpay,
                        'type': 'binary',
                        'datas': file_data,
                        'res_model': 'hr.payslip.run',
                        'res_id': rec.id,  # Attach to the current record
                    })

                    # Return the file as a downloadable attachment
                    return {
                        'type': 'ir.actions.act_url',
                        'url': '/web/content/%s?download=true' % attachment.id,
                        'target': 'self',
                    }
                else:
                    msg = _('No Payslips to process!')
                    raise ValidationError(msg)



    def get_housing_levy(self):
            for rec in self:
                if rec.slip_ids:
                    filename_netpay = 'HOUSING LEVY-' + re.sub(
                        '[^A-Za-z0-9]+', '',
                        rec.name) + '_' + fields.Datetime.context_timestamp(
                            self, fields.Datetime.now()).strftime(
                                '%Y_%m_%d-%H%M%S') + '.xlsx'

                    wb = openpyxl.Workbook()
                    ws = wb.active

                    fr = 7  # First row of data

                    ws['A1'] = self.env.company.name
                    ws['A2'] = 'HOUSING LEVY REPORT'
                    ws['A3'] = 'Month'
                    ws['B3'] = rec.name
                    
                    cols = [
                        'ID NUMBER', 'EMPLOYEE NAME', 'KRA PIN',
                        'HOUSING LEVY AMOUNT'
                    ]


                    # DATA HEADERS
                    for k, x in enumerate(string.ascii_uppercase[0:4]):  # 'ABCD'
                        ws[x + str(fr - 1)] = cols[k]

                    for key, slip in enumerate(rec.slip_ids):
                        ws['A' + str(fr + key)] = slip.employee_id.identification_id or None

                        ws['B' + str(fr + key)] = slip.employee_id.name or None

                        ws['C' + str(
                            fr + key
                        )] = slip.employee_id.kra_pin or None


                        helb = slip.line_ids.search(
                            [('salary_rule_id.code', '=',
                            'HOUSING'),
                            ('slip_id', '=', slip.id)],
                            limit=1).total  # PAYE

                        ws['D' + str(fr + key)] = helb
                            
                    # Totals
                    t = fr + key + 2  # last row for Totals

                    ws['C' + str(t)] = 'TOTAL'
                    # Sum using excel 'SUM' formula
                    ws['D' + str(t)] = '=SUM(D' + str(fr) + ':D' + str(t - 1) + ')'

                    # Ensure the directory exists
                    xls_path = "/tmp/xlsreport.xlsx"


                    if not os.path.exists(xls_path):
                        # Create an empty file
                        with open(xls_path, 'wb') as temp_file:
                            temp_file.write(b'')

                    # Open the file and write the Excel data
                    with open(xls_path, 'wb') as f:
                        wb.save(f)  # Save workbook to file

                    # Read the generated file from the path
                    with open(xls_path, 'rb') as f:
                        file_data = base64.b64encode(f.read())

                    # Create an attachment
                    attachment = self.env['ir.attachment'].create({
                        'name': filename_netpay,
                        'type': 'binary',
                        'datas': file_data,
                        'res_model': 'hr.payslip.run',
                        'res_id': rec.id,  # Attach to the current record
                    })

                    # Return the file as a downloadable attachment
                    return {
                        'type': 'ir.actions.act_url',
                        'url': '/web/content/%s?download=true' % attachment.id,
                        'target': 'self',
                    }
                else:
                    msg = _('No Payslips to process!')
                    raise ValidationError(msg)



    def get_nita_levy(self):
            for rec in self:
                if rec.slip_ids:
                    filename_netpay = 'NITA LEVY-' + re.sub(
                        '[^A-Za-z0-9]+', '',
                        rec.name) + '_' + fields.Datetime.context_timestamp(
                            self, fields.Datetime.now()).strftime(
                                '%Y_%m_%d-%H%M%S') + '.xlsx'

                    wb = openpyxl.Workbook()
                    ws = wb.active

                    fr = 7  # First row of data

                    ws['A1'] = self.env.company.name
                    ws['A2'] = 'NITA LEVY REPORT'
                    ws['A3'] = 'Month'
                    ws['B3'] = rec.name
                    
                    cols = [
                        'ID NUMBER', 'EMPLOYEE NAME', 'KRA PIN',
                        'NITA LEVY AMOUNT'
                    ]


                    # DATA HEADERS
                    for k, x in enumerate(string.ascii_uppercase[0:4]):  # 'ABCD'
                        ws[x + str(fr - 1)] = cols[k]

                    for key, slip in enumerate(rec.slip_ids):
                        ws['A' + str(fr + key)] = slip.employee_id.identification_id or None

                        ws['B' + str(fr + key)] = slip.employee_id.name or None

                        ws['C' + str(
                            fr + key
                        )] = slip.employee_id.kra_pin or None


                        nita = 50  # PAYE

                        ws['D' + str(fr + key)] = nita
                            
                    # Totals
                    t = fr + key + 2  # last row for Totals

                    ws['C' + str(t)] = 'TOTAL'
                    # Sum using excel 'SUM' formula
                    ws['D' + str(t)] = '=SUM(D' + str(fr) + ':D' + str(t - 1) + ')'

                    # Ensure the directory exists
                    xls_path = "/tmp/xlsreport.xlsx"


                    if not os.path.exists(xls_path):
                        # Create an empty file
                        with open(xls_path, 'wb') as temp_file:
                            temp_file.write(b'')

                    # Open the file and write the Excel data
                    with open(xls_path, 'wb') as f:
                        wb.save(f)  # Save workbook to file

                    # Read the generated file from the path
                    with open(xls_path, 'rb') as f:
                        file_data = base64.b64encode(f.read())

                    # Create an attachment
                    attachment = self.env['ir.attachment'].create({
                        'name': filename_netpay,
                        'type': 'binary',
                        'datas': file_data,
                        'res_model': 'hr.payslip.run',
                        'res_id': rec.id,  # Attach to the current record
                    })

                    # Return the file as a downloadable attachment
                    return {
                        'type': 'ir.actions.act_url',
                        'url': '/web/content/%s?download=true' % attachment.id,
                        'target': 'self',
                    }
                else:
                    msg = _('No Payslips to process!')
                    raise ValidationError(msg)

